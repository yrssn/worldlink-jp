"""导入用表格解析工具：支持 CSV / XLSX，统一返回表头与数据行。

- CSV：自动尝试 utf-8-sig / utf-8 / gbk 解码，自动嗅探分隔符
- XLSX：使用 openpyxl 读取首个工作表
- 表头取首行非空内容，缺失列名用 ``列{n}`` 兜底
- 单元格统一转为字符串并去除首尾空白
"""
from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any


class SpreadsheetParseError(ValueError):
    """表格解析失败（格式不支持或内容损坏）。"""


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        # 避免 openpyxl 把整数读成 1.0
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _parse_csv(content: bytes) -> list[list[str]]:
    text = _decode_csv(content)
    if not text.strip():
        return []
    sample = text[:4096]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [[_cell_to_text(cell) for cell in row] for row in reader]


def _parse_xlsx(content: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - 依赖缺失时给出明确提示
        raise SpreadsheetParseError("缺少 openpyxl 依赖，无法解析 Excel 文件") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl 内部异常类型较杂
        raise SpreadsheetParseError("Excel 文件解析失败，请确认文件未损坏") from exc
    try:
        sheet = workbook.active
        if sheet is None:
            return []
        rows: list[list[str]] = []
        for raw_row in sheet.iter_rows(values_only=True):
            rows.append([_cell_to_text(cell) for cell in raw_row])
        return rows
    finally:
        workbook.close()


def _trim_trailing_empty(rows: list[list[str]]) -> list[list[str]]:
    while rows and not any(cell.strip() for cell in rows[-1]):
        rows.pop()
    return rows


def parse_table(filename: str | None, content: bytes) -> tuple[list[str], list[list[str]]]:
    """解析上传文件，返回 ``(headers, data_rows)``。

    - headers：首个非空行作为表头
    - data_rows：表头之后的数据行，每行按表头列数对齐（缺则补空串）
    """
    suffix = Path(filename or "").suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        rows = _parse_xlsx(content)
    elif suffix in (".csv", ".txt", ""):
        rows = _parse_csv(content)
    else:
        raise SpreadsheetParseError("仅支持 CSV 或 XLSX 文件")

    rows = _trim_trailing_empty(rows)
    if not rows:
        return [], []

    header_index = 0
    for index, row in enumerate(rows):
        if any(cell.strip() for cell in row):
            header_index = index
            break

    raw_headers = rows[header_index]
    headers: list[str] = []
    for col_index, name in enumerate(raw_headers):
        clean = name.strip()
        headers.append(clean or f"列{col_index + 1}")

    width = len(headers)
    data_rows: list[list[str]] = []
    for row in rows[header_index + 1:]:
        if not any(cell.strip() for cell in row):
            continue
        normalized = list(row[:width])
        if len(normalized) < width:
            normalized.extend([""] * (width - len(normalized)))
        data_rows.append(normalized)

    return headers, data_rows
