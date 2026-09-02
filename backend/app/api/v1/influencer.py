from __future__ import annotations

import csv
import io
import re
import threading
from datetime import datetime
from pathlib import Path
from typing import Callable

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse
from loguru import logger
from sqlalchemy.orm import Session, joinedload

from app.core.config import settings
from app.core.deps import get_current_user, get_db, is_admin
from app.db.session import SessionLocal
from app.models.bitbrowser import BitBrowserPlatform
from app.models.country import Country
from app.models.dm import DmOutreachLog
from app.models.influencer import Influencer, InfluencerSource, InfluencerStatus
from app.models.influencer_scrape_task import InfluencerScrapeTask
from app.models.post import Post
from app.models.social_account import InfluencerSocialAccount, SocialPlatform
from app.models.user import User
from app.schemas.common import Msg, Page
from app.schemas.dm import DmOutreachLogOut
from app.schemas.influencer import (
    AvatarCacheResult,
    ImportColumnPreview,
    ImportFieldOut,
    ImportPreviewOut,
    ImportResultOut,
    InfluencerCreate,
    InfluencerDetailOut,
    InfluencerFromScrapeRequest,
    InfluencerOut,
    InfluencerScrapeBatchActionResult,
    InfluencerScrapeBatchCreate,
    InfluencerScrapeBatchOut,
    InfluencerScrapeBatchRunRequest,
    InfluencerScrapeRunRequest,
    InfluencerScrapeSaveResult,
    InfluencerScrapeTaskCreate,
    InfluencerScrapeTaskOut,
    InfluencerScrapeTaskSaveRequest,
    InfluencerScrapeTaskUpdate,
    InfluencerUpdate,
    PlatformDetectItem,
    PlatformDetectRequest,
    PlatformOption,
    SocialAccountCreate,
    SocialAccountOut,
    SocialAccountScrapeRequest,
    SocialAccountUpdate,
)
from app.services import apify_service, avatar_cache, influencer_import, influencer_service
from app.utils.csv_export import build_csv, csv_response
from app.utils.platform_detect import (
    KNOWN_PLATFORMS,
    SCRAPABLE_PLATFORMS,
    canonical_platform,
    detect_platform,
    match_platform_code,
)

router = APIRouter(prefix="/influencers", tags=["influencer"])


def _ensure_platform_access(db: Session, owner_id: int, platform_id: int | None) -> None:
    if platform_id is None:
        return
    exists = (
        db.query(BitBrowserPlatform.id)
        .filter(BitBrowserPlatform.id == platform_id)
        .first()
    )
    if not exists:
        raise HTTPException(status_code=400, detail="达人类型不存在")


def _country_or_400(db: Session, country_id: int | None) -> Country | None:
    if country_id is None:
        return None
    row = db.query(Country).filter(Country.id == country_id).first()
    if not row:
        raise HTTPException(status_code=400, detail="国家不存在")
    return row


def _scrape_task_out(db: Session, task: InfluencerScrapeTask) -> InfluencerScrapeTaskOut:
    """把任务序列化为输出，并补上「该主页是否已入库达人」的 influencer_id。"""
    out = InfluencerScrapeTaskOut.model_validate(task)
    result = task.result if isinstance(task.result, dict) else None
    if result:
        if (task.platform or "facebook") == "instagram":
            existing = influencer_service.find_duplicate_social(
                db,
                owner_id=task.owner_id,
                platform=SocialPlatform.instagram,
                handle=result.get("ig_username"),
                url=result.get("ig_url"),
            )
        else:
            existing = influencer_service.find_duplicate(
                db,
                owner_id=task.owner_id,
                fb_page_id=result.get("fb_page_id"),
                fb_page_url=result.get("fb_page_url"),
                email=result.get("email"),
            )
        out.influencer_id = existing.id if existing else task.influencer_id
    return out


def _platform_id_by_code(db: Session) -> dict[str, int]:
    """「平台管理」里的 {平台代码/名称小写: id}，用于按链接识别结果关联平台。

    以「代码」为主（用户可自行改代码来决定自动识别对齐到哪条平台），
    代码没填时退化用平台名称，避免新建平台忘了填代码就识别不出来。
    """
    return influencer_service.platform_id_by_code(db)


def _bind_platform_by_name(db: Session, inf: Influencer, platform: str | None) -> None:
    """按平台规范名（facebook/instagram/...）给达人补上关联平台，已关联则不动。"""
    if inf.platform_id is not None or not platform:
        return
    pid = match_platform_code(platform, _platform_id_by_code(db))
    if pid is None:
        return
    inf.platform_id = pid
    db.commit()


def _default_batch_name(db: Session, owner_id: int) -> str:
    """不传批次名时自动生成：``YYYYMMDD-第N批``（按当天已有批次数递增）。"""
    prefix = datetime.now().strftime("%Y%m%d")
    used = {
        row[0]
        for row in db.query(InfluencerScrapeTask.batch)
        .filter(
            InfluencerScrapeTask.owner_id == owner_id,
            InfluencerScrapeTask.batch.like(f"{prefix}-%"),
        )
        .distinct()
        .all()
    }
    seq = 1
    while f"{prefix}-第{seq}批" in used:
        seq += 1
    return f"{prefix}-第{seq}批"


def _fail_task(db: Session, task: InfluencerScrapeTask, msg: str) -> None:
    task.status = "failed"
    task.error = msg
    task.finished_at = datetime.utcnow()
    db.commit()


def _parse_save_status(value: str | None) -> InfluencerStatus | None:
    """把前端传来的建联状态字符串转成枚举，非法值直接 400。"""
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        return InfluencerStatus(raw)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"不支持的建联状态：{raw}")


def _auto_save_from_task(
    db: Session, task: InfluencerScrapeTask, save_status: str | None = None
) -> None:
    """抓取完成后自动把结果存入建联达人库（发私信后 / 勾选「抓完直接入库」时用）。

    任务来自表格导入 / 列表一键抓取（已绑定 influencer_id）时改为把资料补写回那条达人，不再新建；
    绑定了 social_account_id 时账号维度的结果直接回写到这条关联账号。
    """
    if not isinstance(task.result, dict) or not task.result:
        return
    try:
        if task.influencer_id:
            inf = db.get(Influencer, task.influencer_id)
            if inf is None:
                return
            influencer_service.enrich_influencer_from_form(
                db, inf, task.result, social_account_id=task.social_account_id
            )
            if save_status:
                inf.status = InfluencerStatus(save_status)
                db.commit()
            _bind_platform_by_name(db, inf, task.platform)
            logger.info("[InfluencerScrape task#{}] 已回填达人 #{}", task.id, inf.id)
            return
        if (task.platform or "facebook") == "instagram":
            inf, created = influencer_service.create_influencer_from_ig_form(
                db, owner_id=task.owner_id, form=task.result
            )
        else:
            inf, created = influencer_service.create_influencer_from_form(
                db, owner_id=task.owner_id, form=task.result
            )
        if created and save_status:
            inf.status = InfluencerStatus(save_status)
            db.commit()
        _bind_platform_by_name(db, inf, task.platform)
        logger.info("[InfluencerScrape task#{}] 已自动存入达人库", task.id)
    except Exception as e:  # noqa: BLE001
        logger.warning("[InfluencerScrape task#{}] 自动入库失败：{}", task.id, e)


def _run_scrape_profile_bg(
    task_id: int, auto_save: bool = False, save_status: str | None = None
) -> None:
    """后台线程：按平台跑对应 Apify actor 抓主页资料，映射成可填充表单字段。

    facebook：facebook-pages-scraper；instagram：instagram-profile-scraper。
    auto_save=True 时（如发私信后触发）抓完自动存入达人库。
    """
    db = SessionLocal()
    try:
        task: InfluencerScrapeTask | None = db.get(InfluencerScrapeTask, task_id)
        if not task:
            return
        task.status = "running"
        task.started_at = datetime.utcnow()
        db.commit()

        platform = (task.platform or "facebook").lower()
        if platform not in SCRAPABLE_PLATFORMS:
            _fail_task(db, task, f"平台 {platform} 暂不支持自动抓取资料，请手工录入")
            return
        if platform == "instagram":
            username = apify_service.normalize_ig_username(task.url)
            if not username:
                _fail_task(db, task, "未识别到有效的 Instagram 用户名或主页链接")
                return
            result = apify_service.run_ig_profile([username], db=db)
            items = result.get("items") or []
            if not items or not isinstance(items[0], dict):
                _fail_task(db, task, "未抓取到 Instagram 主页资料，请确认用户名/链接是否有效")
                return
            task.result = avatar_cache.localize_form_avatar(
                influencer_service.ig_profile_to_form(items[0])
            )
            task.status = "done"
            task.finished_at = datetime.utcnow()
            db.commit()
            if auto_save:
                _auto_save_from_task(db, task, save_status)
            logger.info("[InfluencerScrape task#{}] IG done for {}", task_id, task.url)
            return

        scrape_url = influencer_service.normalize_fb_profile_url(task.url)
        result = apify_service.run_fb_pages([scrape_url], max_items=1, db=db)
        items = result.get("items") or []
        if not items or not isinstance(items[0], dict):
            _fail_task(db, task, "未抓取到主页资料，请确认链接是否为有效的 Facebook 主页")
            return

        form = influencer_service.page_profile_to_form(items[0])

        # 个人/创作者主页（Profile）pages-scraper 抓到的资料很稀疏（拿不到昵称/粉丝），
        # 自动用 facebook-profile-scraper 兜底再抓一次并补齐字段。
        if (
            settings.apify_fb_profile_fallback
            and settings.apify_fb_profile_actor
            and influencer_service.fb_form_is_sparse(form)
        ):
            try:
                logger.info(
                    "[InfluencerScrape task#{}] pages 资料稀疏，用 profile actor 兜底：{}",
                    task_id, scrape_url,
                )
                pr = apify_service.run_fb_profile([scrape_url], db=db)
                pitems = pr.get("items") or []
                if pitems and isinstance(pitems[0], dict):
                    fb_form = influencer_service.fb_profile_to_form(pitems[0])
                    if fb_form:
                        form = influencer_service.merge_fb_forms(form, fb_form)
            except Exception as e:  # noqa: BLE001
                logger.warning(
                    "[InfluencerScrape task#{}] profile 兜底抓取失败（保留 pages 结果）：{}",
                    task_id, e,
                )

        # 表单回填的主页链接统一用规整后的标准链接，避免回填群组上下文链接
        form["fb_page_url"] = influencer_service.normalize_fb_profile_url(
            form.get("fb_page_url") or scrape_url
        )
        task.result = avatar_cache.localize_form_avatar(form)
        task.status = "done"
        task.finished_at = datetime.utcnow()
        db.commit()
        if auto_save:
            _auto_save_from_task(db, task, save_status)
        logger.info("[InfluencerScrape task#{}] done for {}", task_id, task.url)
    except Exception as e:  # noqa: BLE001
        logger.exception("[InfluencerScrape task#{}] failed: {}", task_id, e)
        try:
            task = db.get(InfluencerScrapeTask, task_id)
            if task:
                task.status = "failed"
                task.error = str(e)[:2000]
                task.finished_at = datetime.utcnow()
                db.commit()
        except Exception:  # noqa: BLE001
            pass
    finally:
        db.close()


INFLUENCER_CSV_COLUMNS = [
    ("ID", "id"),
    ("名称", "display_name"),
    ("真实姓名", "real_name"),
    ("简介", "bio"),
    ("国家", lambda r: r.country_name or r.country or ""),
    ("国家（英文）", lambda r: r.country_name_en or ""),
    ("国家代码", lambda r: r.country_code or ""),
    ("地区", "region"),
    ("城市", "city"),
    ("语言", "language"),
    ("地址", "address"),
    ("邮箱", "email"),
    ("电话", "phone"),
    ("网站", "website"),
    ("粉丝数", lambda r: getattr(r, "followers", None) or ""),
    (
        "关联账号",
        lambda r: " | ".join(
            f"{a.platform.value if hasattr(a.platform, 'value') else a.platform}:"
            f"{a.title or a.handle or a.url or ''}"
            f"({a.followers if a.followers is not None else '-'})"
            for a in (getattr(r, "accounts", None) or [])
        ),
    ),
    ("账号链接", lambda r: " | ".join(a.url for a in (getattr(r, "accounts", None) or []) if a.url)),
    ("Messenger", lambda r: " | ".join(a.messenger for a in (getattr(r, "accounts", None) or []) if a.messenger)),
    ("账号点赞", lambda r: " | ".join(str(a.likes) for a in (getattr(r, "accounts", None) or []) if a.likes is not None)),
    ("账号评分", lambda r: " | ".join(str(a.rating) for a in (getattr(r, "accounts", None) or []) if a.rating is not None)),
    ("账号备注", lambda r: " | ".join(a.notes for a in (getattr(r, "accounts", None) or []) if a.notes)),
    ("状态", lambda r: r.status.value if r.status else ""),
    ("来源", lambda r: r.source.value if r.source else ""),
    ("类型", lambda r: r.platform_name or ""),
    ("建联进度", "progress"),
    ("建联用户", lambda r: r.owner_name or ""),
    ("KOL 编号", "code"),
    ("KOL 公司名", "company"),
    ("性别", "gender"),
    ("建联负责人", "contact_owner"),
    ("落地负责人", "landing_owner"),
    ("来源渠道", "source_channel"),
    ("建联开始日期", "contact_started_at"),
    ("计划首次来访时间", "planned_visit_at"),
    ("是否推特", lambda r: "" if r.has_twitter is None else ("是" if r.has_twitter else "否")),
    ("推特渠道", "twitter_channel"),
    ("群名称", "group_name"),
    ("标签", "tags"),
    ("备注", "notes"),
    ("创建时间", "created_at"),
]


def _followers_expr():
    """达人粉丝数口径：各关联账号粉丝数取最大值；没有任何账号时才退化用旧的主表 fb_followers。"""
    from sqlalchemy import func, select

    max_account = (
        select(func.max(InfluencerSocialAccount.followers))
        .where(InfluencerSocialAccount.influencer_id == Influencer.id)
        .correlate(Influencer)
        .scalar_subquery()
    )
    return func.coalesce(max_account, Influencer.fb_followers, 0)


def _account_keyword_exists(like: str):
    """关键词命中任一关联账号的链接 / handle / 账号名。"""
    from sqlalchemy import exists

    return exists().where(
        InfluencerSocialAccount.influencer_id == Influencer.id,
        (InfluencerSocialAccount.url.like(like))
        | (InfluencerSocialAccount.handle.like(like))
        | (InfluencerSocialAccount.title.like(like)),
    )


def _apply_influencer_filters(
    q,
    keyword: str | None,
    status_eq: str | None,
    country: str | None,
    platform_id: int | None,
    country_id: int | None = None,
    followers_min: int | None = None,
    followers_max: int | None = None,
):
    """达人列表/导出共用的过滤条件（关键词 / 状态 / 国家 / 关联平台 / 粉丝区间）。"""
    if keyword:
        like = f"%{keyword}%"
        q = q.filter(
            (Influencer.display_name.like(like))
            | (Influencer.real_name.like(like))
            | (Influencer.email.like(like))
            | (Influencer.fb_page_url.like(like))
            | _account_keyword_exists(like)
        )
    if status_eq:
        q = q.filter(Influencer.status == status_eq)
    if country:
        c = country.strip()
        if c == "__none__":
            q = q.filter(Influencer.country.is_(None))
        elif c:
            q = q.filter(Influencer.country == c)
    if country_id is not None:
        if country_id == 0:
            q = q.filter(Influencer.country_id.is_(None))
        else:
            q = q.filter(Influencer.country_id == country_id)
    if platform_id is not None:
        if platform_id == 0:
            q = q.filter(Influencer.platform_id.is_(None))
        else:
            q = q.filter(Influencer.platform_id == platform_id)
    if followers_min is not None or followers_max is not None:
        followers = _followers_expr()
        if followers_min is not None:
            q = q.filter(followers >= followers_min)
        if followers_max is not None:
            q = q.filter(followers <= followers_max)
    return q


@router.get("", response_model=Page[InfluencerOut])
def list_influencers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    keyword: str | None = None,
    status_eq: str | None = Query(None, alias="status"),
    country: str | None = Query(None, description="国家代码（兼容旧参数），__none__ = 未填"),
    country_id: int | None = Query(None, description="关联国家 id，0 = 未关联"),
    platform_id: int | None = Query(None, description="关联平台 id，0 = 未关联"),
    followers_min: int | None = Query(None, ge=0, description="粉丝数下限"),
    followers_max: int | None = Query(None, ge=0, description="粉丝数上限"),
    sort: str = Query("id_desc", description="id_desc / followers_desc / followers_asc"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(Influencer).options(
        joinedload(Influencer.platform),
        joinedload(Influencer.country_ref),
        joinedload(Influencer.owner),
    )
    if not is_admin(user):
        q = q.filter(Influencer.owner_id == user.id)
    q = _apply_influencer_filters(
        q, keyword, status_eq, country, platform_id, country_id,
        followers_min, followers_max,
    )
    total = q.count()
    if sort == "followers_desc":
        q = q.order_by(_followers_expr().desc(), Influencer.id.desc())
    elif sort == "followers_asc":
        q = q.order_by(_followers_expr().asc(), Influencer.id.desc())
    else:
        q = q.order_by(Influencer.id.desc())
    items = q.offset((page - 1) * page_size).limit(page_size).all()
    _mark_has_outreach(db, items)
    _attach_accounts(db, items)
    return Page[InfluencerOut](total=total, page=page, page_size=page_size, items=items)


def _attach_accounts(db: Session, items: list[Influencer]) -> None:
    """给列表里的达人挂上关联社交账号，并算出粉丝数（各账号取最大；没账号时退化用旧 fb_followers）。"""
    ids = [i.id for i in items]
    if not ids:
        return
    grouped: dict[int, list[SocialAccountOut]] = {}
    rows = (
        db.query(InfluencerSocialAccount)
        .filter(InfluencerSocialAccount.influencer_id.in_(ids))
        .order_by(InfluencerSocialAccount.id.asc())
        .all()
    )
    for row in rows:
        grouped.setdefault(row.influencer_id, []).append(
            SocialAccountOut.model_validate(row)
        )
    for i in items:
        i.accounts = grouped.get(i.id, [])
        counts = [a.followers for a in i.accounts if a.followers is not None]
        if not i.accounts and i.fb_followers is not None:
            counts.append(i.fb_followers)
        i.followers = max(counts) if counts else None


def _mark_has_outreach(db: Session, items: list[Influencer]) -> None:
    """给列表里的达人标注 has_outreach（是否已私信过）。"""
    ids = [i.id for i in items]
    if not ids:
        return
    contacted = {
        row[0]
        for row in db.query(DmOutreachLog.influencer_id)
        .filter(DmOutreachLog.influencer_id.in_(ids))
        .distinct()
        .all()
    }
    for i in items:
        i.has_outreach = i.id in contacted


@router.get("/export")
def export_influencers(
    keyword: str | None = None,
    status_eq: str | None = Query(None, alias="status"),
    country: str | None = None,
    country_id: int | None = None,
    platform_id: int | None = None,
    followers_min: int | None = Query(None, ge=0),
    followers_max: int | None = Query(None, ge=0),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """按当前过滤条件导出达人列表（CSV）。"""
    q = db.query(Influencer).options(
        joinedload(Influencer.platform),
        joinedload(Influencer.country_ref),
        joinedload(Influencer.owner),
    )
    if not is_admin(user):
        q = q.filter(Influencer.owner_id == user.id)
    q = _apply_influencer_filters(
        q, keyword, status_eq, country, platform_id, country_id,
        followers_min, followers_max,
    )
    rows = q.order_by(Influencer.id.desc()).all()
    _attach_accounts(db, rows)
    data = build_csv(rows, INFLUENCER_CSV_COLUMNS)
    return csv_response("influencers.csv", data)


@router.get("/countries", response_model=list[str])
def list_influencer_countries(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """当前可见达人里已用过的国家分类（供前端筛选/补全）。"""
    q = db.query(Influencer.country).distinct()
    if not is_admin(user):
        q = q.filter(Influencer.owner_id == user.id)
    return sorted({(row[0] or "").strip() for row in q.all() if (row[0] or "").strip()})


@router.get("/platform-options", response_model=list[PlatformOption])
def list_platform_options(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """可选平台：取「平台管理」里的记录，按代码/名称对回平台规范名。

    拓展平台只要在「平台管理」新建并填上代码就会出现在这里；
    平台管理里一条都没对上时退回内置平台名，避免下拉空掉。
    """
    rows = (
        db.query(BitBrowserPlatform)
        .order_by(BitBrowserPlatform.sort_order.asc(), BitBrowserPlatform.id.asc())
        .all()
    )
    options: list[PlatformOption] = []
    seen: set[str] = set()
    for row in rows:
        platform = canonical_platform(row.code, row.name)
        if not platform or platform in seen:
            continue
        seen.add(platform)
        options.append(
            PlatformOption(
                platform=platform,
                name=row.name,
                platform_id=row.id,
                code=row.code,
                scrapable=platform in SCRAPABLE_PLATFORMS,
            )
        )
    if not options:
        options = [
            PlatformOption(
                platform=name,
                name=name,
                scrapable=name in SCRAPABLE_PLATFORMS,
            )
            for name in KNOWN_PLATFORMS
        ]
    return options


@router.get("/avatars/{filename}")
def get_influencer_avatar(filename: str):
    """返回已缓存到本机的达人头像（文件名是原地址 sha1，供 <img> 直接加载）。"""
    path = avatar_cache.avatar_path(filename)
    if path is None or not path.is_file():
        raise HTTPException(status_code=404, detail="头像不存在")
    return FileResponse(path)


@router.post("/avatars/cache", response_model=AvatarCacheResult)
def cache_influencer_avatars(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    limit: int = Query(200, ge=1, le=1000, description="本次最多处理多少条"),
):
    """把存量达人的远端头像下载到本机（国内免代理看图）。"""
    q = db.query(Influencer).filter(Influencer.avatar_url.like("http%"))
    if not is_admin(user):
        q = q.filter(Influencer.owner_id == user.id)
    rows = q.order_by(Influencer.id.desc()).limit(limit).all()
    cached = 0
    failed = 0
    for row in rows:
        local = avatar_cache.localize_avatar(row.avatar_url)
        if avatar_cache.is_local_avatar(local):
            row.avatar_url = local
            cached += 1
        else:
            failed += 1
    db.commit()
    return AvatarCacheResult(total=len(rows), cached=cached, failed=failed)


@router.post("/detect-platforms", response_model=list[PlatformDetectItem])
def detect_platforms(
    payload: PlatformDetectRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """按链接正则预分类：返回每条链接识别到的平台，并对齐「平台管理」里的平台代码。"""
    codes = _platform_id_by_code(db)
    names = {
        pid: name
        for pid, name in db.query(BitBrowserPlatform.id, BitBrowserPlatform.name).all()
    }
    items: list[PlatformDetectItem] = []
    for raw in payload.urls or []:
        url = (raw or "").strip()
        if not url:
            continue
        platform = detect_platform(url)
        pid = match_platform_code(platform, codes) if platform else None
        items.append(
            PlatformDetectItem(
                url=url,
                platform=platform,
                platform_id=pid,
                platform_name=names.get(pid) if pid else None,
                scrapable=bool(platform and platform in SCRAPABLE_PLATFORMS),
            )
        )
    return items


@router.post("", response_model=InfluencerOut)
def create_influencer(
    payload: InfluencerCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    data = payload.model_dump(exclude={"social_accounts"})
    _ensure_platform_access(db, user.id, data.get("platform_id"))
    country = _country_or_400(db, data.get("country_id"))
    if country is not None:
        data["country"] = country.code
    inf = Influencer(**data, owner_id=user.id)
    db.add(inf)
    db.flush()
    for sa in payload.social_accounts or []:
        fields = sa.model_dump()
        if fields.get("platform_id") is None:
            fields["platform_id"] = influencer_service.resolve_platform_id(db, sa.platform)
        if not fields.get("handle"):
            fields["handle"] = influencer_service.handle_from_url(
                fields.get("url"), fields.get("page_id")
            )
        db.add(InfluencerSocialAccount(influencer_id=inf.id, **fields))
    db.commit()
    db.refresh(inf)
    return inf


@router.post("/scrape-profile", response_model=InfluencerScrapeTaskOut)
def start_scrape_profile(
    payload: InfluencerScrapeTaskCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """「自动抓取」：按平台异步跑对应 actor（FB 主页 / IG 主页资料）。

    主页抓取较慢，放后台线程执行；前端轮询任务状态，done 后用 result 自动填充表单。
    """
    url = (payload.url or "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="请填写主页链接或用户名")
    platform = (payload.platform or "").strip().lower() or (
        detect_platform(url) or "facebook"
    )
    if platform == "auto":
        platform = detect_platform(url) or "facebook"
    if platform not in SCRAPABLE_PLATFORMS:
        raise HTTPException(status_code=400, detail=f"平台 {platform} 暂不支持自动抓取")
    task = InfluencerScrapeTask(
        owner_id=user.id, platform=platform, url=url, status="pending"
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    threading.Thread(
        target=_run_scrape_profile_bg, args=(task.id,), daemon=True
    ).start()
    return _scrape_task_out(db, task)


#: 上传的链接表格大小上限
_MAX_IMPORT_BYTES = 5 * 1024 * 1024


def _looks_like_profile_value(value: str) -> bool:
    """单元格内容是否像主页链接 / 用户名（用于忽略表头和无关列）。"""
    s = value.strip()
    if not s or len(s) > 500 or " " in s:
        return False
    return s.startswith(("http://", "https://", "www.", "@")) or "." in s


def _extract_urls_from_upload(filename: str, content: bytes) -> list[str]:
    """从上传的 xlsx / csv / txt 里取出所有像主页链接 / 用户名的单元格。"""
    suffix = Path(filename).suffix.lower()
    values: list[str] = []
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    values.extend(str(c) for c in row if isinstance(c, str))
        finally:
            wb.close()
    else:
        text = content.decode("utf-8-sig", errors="ignore")
        if suffix == ".csv" or "," in text or "\t" in text:
            for row in csv.reader(io.StringIO(text), delimiter="\t" if "\t" in text else ","):
                values.extend(row)
        else:
            values.extend(text.splitlines())
    return [v.strip() for v in values if _looks_like_profile_value(v)]


def _stage_urls(
    db: Session,
    user: User,
    raw_urls: list[str],
    platform: str | None,
    fallback_platform: str | None,
    batch: str | None,
) -> list[InfluencerScrapeTaskOut]:
    """把一批链接去重后写入暂存区（status=staged），逐条按链接识别平台。"""
    plat = (platform or "auto").strip().lower()
    if plat != "auto" and plat not in KNOWN_PLATFORMS:
        raise HTTPException(status_code=400, detail="不支持的抓取平台")
    fallback = (fallback_platform or "facebook").strip().lower()
    if fallback not in KNOWN_PLATFORMS:
        raise HTTPException(status_code=400, detail="不支持的兼容平台")
    seen: set[str] = set()
    urls: list[str] = []
    for raw in raw_urls:
        u = (raw or "").strip()
        if not u or u in seen:
            continue
        seen.add(u)
        urls.append(u)
    if not urls:
        raise HTTPException(status_code=400, detail="没有可导入的链接")
    batch_name = (batch or "").strip() or _default_batch_name(db, user.id)
    created: list[InfluencerScrapeTask] = []
    for u in urls:
        row_platform = plat
        if row_platform == "auto":
            row_platform = detect_platform(u) or fallback
        task = InfluencerScrapeTask(
            owner_id=user.id,
            platform=row_platform,
            url=u,
            batch=batch_name,
            status="staged",
        )
        db.add(task)
        created.append(task)
    db.commit()
    for t in created:
        db.refresh(t)
    return [_scrape_task_out(db, t) for t in created]


@router.post("/scrape-profile/batch", response_model=list[InfluencerScrapeTaskOut])
def batch_stage_scrape_profiles(
    payload: InfluencerScrapeBatchCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """批量导入链接到暂存区（status=staged）：只保存不跑抓取，按 平台 + 批次名 分组。

    之后可在暂存列表里逐条/整批「抓取」或「私信建联」，私信发出后自动入库。
    """
    return _stage_urls(
        db,
        user,
        payload.urls or [],
        platform=payload.platform,
        fallback_platform=payload.fallback_platform,
        batch=payload.batch,
    )


@router.post("/scrape-profile/batch/upload", response_model=list[InfluencerScrapeTaskOut])
async def upload_scrape_profile_batch(
    file: UploadFile = File(..., description="主页链接表格：xlsx / csv / txt"),
    platform: str = Form("auto"),
    fallback_platform: str = Form("facebook"),
    url_column: int | None = Form(None, description="主页链接列下标，不传则全表扫描"),
    has_header: bool = Form(True),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """上传主页链接表格（xlsx / csv / txt）批量导入到暂存区。

    指定 ``url_column`` 时只取该列；不指定则表格里所有形如链接 / @用户名 的单元格都会被取出。
    """
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件是空的")
    if len(content) > _MAX_IMPORT_BYTES:
        raise HTTPException(status_code=400, detail="文件过大（上限 5MB）")
    if url_column is not None:
        rows = _read_table(file.filename or "", content)
        if has_header:
            rows = rows[1:]
        urls = [u for u in (_cell(row, url_column) for row in rows) if u]
    else:
        urls = _extract_urls_from_upload(file.filename or "", content)
    if not urls:
        raise HTTPException(status_code=400, detail="表格里没有识别到主页链接 / 用户名")
    return _stage_urls(
        db,
        user,
        urls,
        platform=platform,
        fallback_platform=fallback_platform,
        batch=None,
    )


def _read_table(filename: str, content: bytes) -> list[list[str]]:
    """把上传的 xlsx / csv / txt 读成二维表（首行为表头），单元格统一转字符串。"""
    suffix = Path(filename).suffix.lower()
    rows: list[list[str]] = []
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            for row in ws.iter_rows(values_only=True):
                rows.append(["" if c is None else str(c).strip() for c in row])
        finally:
            wb.close()
    else:
        text = content.decode("utf-8-sig", errors="ignore")
        delimiter = "\t" if ("\t" in text and "," not in text) else ","
        for row in csv.reader(io.StringIO(text), delimiter=delimiter):
            rows.append([(c or "").strip() for c in row])
    return [r for r in rows if any(c for c in r)]


def _cell(row: list[str], index: int | None) -> str:
    """按列下标安全取值，越界/未选列返回空串。"""
    if index is None or index < 0 or index >= len(row):
        return ""
    return (row[index] or "").strip()


def _looks_like_url_value(value: str) -> bool:
    """单元格内容是否像主页链接（比 ``_looks_like_profile_value`` 严格，避免把邮箱/数字列认成链接）。"""
    s = value.strip().lower()
    if not s or " " in s or "@" in s:
        return False
    if s.startswith(("http://", "https://")):
        return True
    return s.startswith("www.") and "/" in s


def _url_variants(url: str) -> list[str]:
    """主页链接的常见等价写法（http/https、带不带 www、带不带尾斜杠），用于查重。"""
    base = influencer_service.normalize_fb_url(url)
    if not base:
        return []
    for prefix in ("https://", "http://"):
        if base.startswith(prefix):
            base = base[len(prefix):]
            break
    hosts = [base]
    if base.startswith("www."):
        hosts.append(base[4:])
    elif base.startswith(("m.", "web.")):
        hosts.append(base.split(".", 1)[1])
    else:
        hosts.append(f"www.{base}")
    return [
        f"{scheme}{h}{suffix}"
        for h in hosts
        for scheme in ("https://", "http://")
        for suffix in ("", "/")
    ] + [url]


def _scrape_task_url_index(db: Session, owner_id: int) -> dict[str, tuple[int, int | None]]:
    """之前导入 / 抓取任务用过的原始链接（归一化） -> 达人 id。

    导入时填的是分享链接（如 facebook.com/share/xxx/），抓取后账号行的 url 会换成规范主页链接，
    再次导入同一张表时只能靠任务表里留下的原始链接对回去。
    """
    rows = (
        db.query(
            InfluencerScrapeTask.url,
            InfluencerScrapeTask.influencer_id,
            InfluencerScrapeTask.social_account_id,
        )
        .filter(
            InfluencerScrapeTask.owner_id == owner_id,
            InfluencerScrapeTask.influencer_id.isnot(None),
        )
        .order_by(InfluencerScrapeTask.id.asc())
        .all()
    )
    index: dict[str, tuple[int, int | None]] = {}
    for url, iid, sid in rows:
        norm = influencer_service.normalize_fb_url(url)
        if norm and iid is not None:
            index[norm] = (iid, sid)
    return index


def _find_by_task_url(
    db: Session, index: dict[str, tuple[int, int | None]], norm_url: str
) -> tuple[Influencer | None, int | None]:
    hit = index.get(norm_url)
    if hit is None:
        return None, None
    iid, sid = hit
    inf = (
        db.query(Influencer)
        .filter(Influencer.id == iid, Influencer.deleted_at.is_(None))
        .first()
    )
    return inf, (sid if inf is not None else None)


def _handle_from_url(url: str) -> str | None:
    """从主页链接里取账号名（路径最后一段），取不到返回 ``None``。"""
    parts = influencer_service.normalize_fb_url(url).rsplit("/", 1)
    if len(parts) < 2:
        return None
    return parts[1].strip().lstrip("@") or None


def _to_followers(value: str) -> int | None:
    """把表格里的粉丝数（可能带逗号 / 万 / k / w 后缀）转成整数。"""
    text = (value or "").strip().lower().replace(",", "").replace(" ", "")
    if not text:
        return None
    multiplier = 1.0
    if text.endswith(("万", "w")):
        multiplier, text = 10000.0, text[:-1]
    elif text.endswith("k"):
        multiplier, text = 1000.0, text[:-1]
    elif text.endswith("m"):
        multiplier, text = 1000000.0, text[:-1]
    try:
        return int(float(text) * multiplier)
    except ValueError:
        return None


def _import_person_fields(
    row: list[str],
    col: Callable[[list[str], str], str],
    countries: dict[str, Country],
) -> dict[str, object]:
    """从一行表格取出落 ``influencers`` 的人/建联维度字段（空值不写）。"""
    out: dict[str, object] = {}
    for key in influencer_import.INFLUENCER_TEXT_FIELDS:
        value = col(row, key)
        if value:
            out[key] = value
    if "gender" in out:
        out["gender"] = influencer_import.parse_gender(str(out["gender"]))
    for key in influencer_import.INFLUENCER_DATE_FIELDS:
        parsed = influencer_import.parse_date(col(row, key))
        if parsed is not None:
            out[key] = parsed
    has_twitter = influencer_import.parse_bool(col(row, "has_twitter"))
    if has_twitter is not None:
        out["has_twitter"] = has_twitter
    country_text = col(row, "country")
    if country_text:
        country = influencer_import.match_country(country_text, countries)
        if country is not None:
            out["country_id"] = country.id
            out["country"] = country.code or country.name_zh
        else:
            out["country"] = country_text[:64]
    return out


@router.post("/import/preview", response_model=ImportPreviewOut)
async def preview_import_table(
    file: UploadFile = File(..., description="存量数据表：xlsx / csv / txt"),
    user: User = Depends(get_current_user),
):
    """导入第一步：回显表头与前几行样例，供前端选“哪一列是主页链接”。"""
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件是空的")
    if len(content) > _MAX_IMPORT_BYTES:
        raise HTTPException(status_code=400, detail="文件过大（上限 5MB）")
    rows = _read_table(file.filename or "", content)
    if not rows:
        raise HTTPException(status_code=400, detail="表格里没有可读取的内容")
    header, body = rows[0], rows[1:]
    width = max(len(r) for r in rows)
    columns: list[ImportColumnPreview] = []
    for idx in range(width):
        samples = [_cell(r, idx) for r in body[:5]]
        samples = [s for s in samples if s]
        hits = sum(1 for r in body[:20] if _looks_like_url_value(_cell(r, idx)))
        columns.append(
            ImportColumnPreview(
                index=idx,
                name=_cell(header, idx) or f"第{idx + 1}列",
                samples=samples,
                looks_like_url=hits > 0,
            )
        )
    suggested = next((c.index for c in columns if c.looks_like_url), None)
    suggested_columns = influencer_import.suggest_columns([_cell(header, i) for i in range(width)])
    if suggested is not None:
        suggested_columns["url"] = suggested
    elif "url" in suggested_columns:
        suggested = suggested_columns["url"]
    return ImportPreviewOut(
        filename=file.filename or "",
        total_rows=len(body),
        columns=columns,
        suggested_url_column=suggested,
        suggested_columns=suggested_columns,
        fields=[
            ImportFieldOut(key=f.key, label=f.label, required=f.required)
            for f in influencer_import.IMPORT_FIELDS
        ],
    )


@router.post("/import", response_model=ImportResultOut)
async def import_influencers(
    file: UploadFile = File(..., description="存量数据表：xlsx / csv / txt"),
    url_column: int = Form(..., description="主页链接列下标"),
    name_column: int | None = Form(None, description="昵称列下标"),
    email_column: int | None = Form(None, description="邮箱列下标"),
    followers_column: int | None = Form(None, description="粉丝数列下标"),
    notes_column: int | None = Form(None, description="备注列下标"),
    column_map: str | None = Form(
        None,
        description='其余表头映射 JSON：{"code": 1, "company": 4, ...}，键见 /import/preview 的 fields',
    ),
    has_header: bool = Form(True, description="首行是否为表头"),
    status: str = Form("pre_contact", description="导入后的建联状态"),
    scrape: bool = Form(False, description="是否对主页链接抓取内容"),
    platform: str = Form("auto", description="auto = 按链接自动识别平台"),
    fallback_platform: str = Form("other", description="识别不出平台的链接归为哪个平台"),
    batch: str | None = Form(None, description="抓取批次名，不传自动生成"),
    create_missing: bool = Form(True, description="链接匹配不到已有达人时是否新建；False = 只补已有达人的资料"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """按选定的主页链接列导入存量数据：逐行建达人，可选建联状态与是否抓取。

    已有达人按链接唯一匹配（关联账号链接、旧主表 fb_page_url、以及之前导入/抓取任务用过的原始链接，
    归一化比对），命中则只补空字段不新建；``create_missing=False`` 时匹配不到的行直接跳过。

    scrape=True 时同时建好抓取任务并后台开跑，抓到的资料回填到对应达人（不重复建档）；
    暂不支持自动抓资料的平台（如 TikTok / 小红书）只建暂存任务，不会报错中断。
    """
    save_status = _parse_save_status(status) or InfluencerStatus.pre_contact
    plat = (platform or "auto").strip().lower()
    if plat != "auto" and plat not in KNOWN_PLATFORMS:
        raise HTTPException(status_code=400, detail="不支持的抓取平台")
    fallback = (fallback_platform or "other").strip().lower()
    if fallback not in KNOWN_PLATFORMS and fallback != "other":
        raise HTTPException(status_code=400, detail="不支持的兼容平台")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件是空的")
    if len(content) > _MAX_IMPORT_BYTES:
        raise HTTPException(status_code=400, detail="文件过大（上限 5MB）")
    rows = _read_table(file.filename or "", content)
    header_row: list[str] = rows[0] if (has_header and rows) else []
    if has_header:
        rows = rows[1:]
    if not rows:
        raise HTTPException(status_code=400, detail="表格里没有可导入的数据行")

    batch_name = (batch or "").strip() or _default_batch_name(db, user.id)
    platform_codes = _platform_id_by_code(db)
    countries = influencer_import.country_lookup(db)
    cols = influencer_import.parse_column_map(column_map)
    for key, idx in (
        ("url", url_column),
        ("name", name_column),
        ("email", email_column),
        ("followers", followers_column),
        ("notes", notes_column),
    ):
        if idx is not None:
            cols[key] = idx

    def col(row: list[str], key: str) -> str:
        idx = cols.get(key)
        return _cell(row, idx) if idx is not None else ""

    # 表头写「粉丝数量（K）」时，纯数字按千换算
    followers_in_k = "k" in col(header_row, "followers").lower()

    created = duplicated = skipped = updated = 0
    seen: set[str] = set()
    tasks: list[InfluencerScrapeTask] = []
    task_url_index = _scrape_task_url_index(db, user.id)
    for row in rows:
        url = col(row, "url")
        if not url:
            skipped += 1
            continue
        key = influencer_service.normalize_fb_url(url)
        if key in seen:
            duplicated += 1
            continue
        seen.add(key)
        row_platform = (
            plat
            if plat != "auto"
            else (
                detect_platform(url)
                or canonical_platform(col(row, "platform"))
                or fallback
            )
        )
        social = SocialPlatform(row_platform) if row_platform in {
            p.value for p in SocialPlatform
        } else SocialPlatform.other

        # 查重容忍 http/https、www、尾斜杠、query 等写法差异
        email = col(row, "email") or None
        existing = None
        for variant in _url_variants(url):
            if row_platform == "facebook":
                existing = influencer_service.find_duplicate(
                    db, owner_id=user.id, fb_page_url=variant
                )
            else:
                existing = influencer_service.find_duplicate_social(
                    db, owner_id=user.id, platform=social, url=variant
                )
            if existing:
                break
        account_id: int | None = None
        if existing is None:
            existing, account_id = _find_by_task_url(db, task_url_index, key)
        if existing is not None and account_id is None:
            # 已有达人该平台只有一个账号时直接补到这一行，避免分享链接 vs 规范链接写法不同又建一行
            same = (
                db.query(InfluencerSocialAccount.id)
                .filter(
                    InfluencerSocialAccount.influencer_id == existing.id,
                    InfluencerSocialAccount.platform == social,
                )
                .limit(2)
                .all()
            )
            if len(same) == 1:
                account_id = same[0][0]
        if existing is None and email:
            existing = influencer_service.find_duplicate(db, owner_id=user.id, email=email)
        if existing is None and not create_missing:
            skipped += 1
            continue
        person = _import_person_fields(row, col, countries)
        result_status = influencer_import.parse_result_status(col(row, "result"))
        inf = existing
        if inf is None:
            inf = Influencer(
                display_name=col(row, "name") or col(row, "title") or url,
                status=result_status or save_status,
                source=InfluencerSource.manual,
                owner_id=user.id,
                platform_id=match_platform_code(row_platform, platform_codes),
                **person,
            )
            db.add(inf)
            db.flush()
            created += 1
        else:
            # 已有达人只补空值，不覆盖人工维护的数据
            changed = False
            for k, v in person.items():
                if v not in (None, "") and getattr(inf, k) in (None, ""):
                    setattr(inf, k, v)
                    changed = True
            if result_status and inf.status == InfluencerStatus.pre_contact and result_status != inf.status:
                inf.status = result_status
                changed = True
            if changed:
                updated += 1
            else:
                duplicated += 1
        followers = _to_followers(col(row, "followers"))
        if followers is not None and followers_in_k and not re.search(
            r"[a-z万]", col(row, "followers").lower()
        ):
            followers *= 1000
        account = influencer_service.upsert_social_account(
            db,
            inf.id,
            social,
            handle=_handle_from_url(url),
            url=url,
            followers=followers,
            keep_existing_url=True,
            platform_id=match_platform_code(row_platform, platform_codes),
            fields={"title": col(row, "title") or None},
            account_id=account_id,
        )
        if scrape:
            scrapable = row_platform in SCRAPABLE_PLATFORMS
            if account is not None and account.id is None:
                db.flush()
            task = InfluencerScrapeTask(
                owner_id=user.id,
                platform=row_platform,
                url=url,
                batch=batch_name,
                status="pending" if scrapable else "staged",
                influencer_id=inf.id,
                social_account_id=account.id if account is not None else None,
            )
            db.add(task)
            if scrapable:
                tasks.append(task)
    db.commit()

    for task in tasks:
        db.refresh(task)
        threading.Thread(
            target=_run_scrape_profile_bg,
            args=(task.id, True, save_status.value),
            daemon=True,
        ).start()
    return ImportResultOut(
        total_rows=len(rows),
        created=created,
        updated=updated,
        duplicated=duplicated,
        skipped=skipped,
        scrape_tasks=len(tasks),
        batch=batch_name if scrape else None,
    )


@router.get("/scrape-profile", response_model=list[InfluencerScrapeTaskOut])
def list_scrape_profiles(
    limit: int = Query(100, ge=1, le=500),
    platform: str | None = None,
    batch: str | None = None,
    status_eq: str | None = Query(None, alias="status"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """抓取/暂存任务列表：按创建时间倒序，支持按 平台 / 批次 / 状态 过滤。"""
    q = db.query(InfluencerScrapeTask)
    if not is_admin(user):
        q = q.filter(InfluencerScrapeTask.owner_id == user.id)
    if platform:
        q = q.filter(InfluencerScrapeTask.platform == platform.strip().lower())
    if batch is not None:
        b = batch.strip()
        if b == "":
            q = q.filter(InfluencerScrapeTask.batch.is_(None))
        else:
            q = q.filter(InfluencerScrapeTask.batch == b)
    if status_eq:
        q = q.filter(InfluencerScrapeTask.status == status_eq)
    tasks = q.order_by(InfluencerScrapeTask.id.desc()).limit(limit).all()
    return [_scrape_task_out(db, t) for t in tasks]


@router.get("/scrape-profile-batches", response_model=list[InfluencerScrapeBatchOut])
def list_scrape_batches(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """批次汇总：按 平台 + 批次名 分组，返回总数/暂存数/已抓取数（供筛选下拉）。"""
    from sqlalchemy import case, func

    q = db.query(
        InfluencerScrapeTask.platform,
        InfluencerScrapeTask.batch,
        InfluencerScrapeTask.owner_id,
        func.count(InfluencerScrapeTask.id),
        func.sum(case((InfluencerScrapeTask.status == "staged", 1), else_=0)),
        func.sum(
            case((InfluencerScrapeTask.status.in_(("pending", "running")), 1), else_=0)
        ),
        func.sum(case((InfluencerScrapeTask.status == "failed", 1), else_=0)),
        func.sum(case((InfluencerScrapeTask.status == "done", 1), else_=0)),
    )
    if not is_admin(user):
        q = q.filter(InfluencerScrapeTask.owner_id == user.id)
    rows = (
        q.group_by(
            InfluencerScrapeTask.platform,
            InfluencerScrapeTask.batch,
            InfluencerScrapeTask.owner_id,
        )
        .order_by(InfluencerScrapeTask.batch.desc(), InfluencerScrapeTask.platform)
        .all()
    )
    owner_ids = {oid for (_p, _b, oid, *_rest) in rows if oid is not None}
    owner_names = {
        uid: name
        for uid, name in db.query(User.id, User.username)
        .filter(User.id.in_(owner_ids))
        .all()
    } if owner_ids else {}
    return [
        InfluencerScrapeBatchOut(
            platform=p,
            batch=b,
            owner_id=oid,
            owner_name=owner_names.get(oid) if oid is not None else None,
            total=int(total or 0),
            staged=int(staged or 0),
            running=int(running or 0),
            failed=int(failed or 0),
            done=int(done or 0),
        )
        for (p, b, oid, total, staged, running, failed, done) in rows
    ]


def _batch_task_query(db: Session, user: User, batch: str | None, platform: str | None):
    """定位当前用户可见的抓取任务：batch=None 为不限批次，"" 为未分组。"""
    q = db.query(InfluencerScrapeTask)
    if not is_admin(user):
        q = q.filter(InfluencerScrapeTask.owner_id == user.id)
    if batch is not None:
        b = batch.strip()
        if b:
            q = q.filter(InfluencerScrapeTask.batch == b)
        else:
            q = q.filter(InfluencerScrapeTask.batch.is_(None))
    if platform:
        q = q.filter(InfluencerScrapeTask.platform == platform.strip().lower())
    return q


@router.post("/scrape-profile-batches/run", response_model=InfluencerScrapeBatchActionResult)
def run_scrape_batch(
    payload: InfluencerScrapeBatchRunRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """批量抓取：把暂存（可选含失败）的任务一次性丢进抓取队列。

    auto_save=True 时抓完直接入库，save_status 指定入库后的建联状态；
    暂不支持自动抓资料的平台（如 TikTok / 小红书）计入 skipped，不会报错中断。
    """
    statuses = ["staged"] + (["failed"] if payload.include_failed else [])
    tasks = (
        _batch_task_query(db, user, payload.batch, payload.platform)
        .filter(InfluencerScrapeTask.status.in_(statuses))
        .all()
    )
    runnable = [t for t in tasks if (t.platform or "facebook") in SCRAPABLE_PLATFORMS]
    status = _parse_save_status(payload.save_status)
    for t in runnable:
        t.status = "pending"
        t.error = None
        t.result = None
        t.started_at = None
        t.finished_at = None
    db.commit()
    for t in runnable:
        threading.Thread(
            target=_run_scrape_profile_bg,
            args=(t.id, payload.auto_save, status.value if status else None),
            daemon=True,
        ).start()
    return InfluencerScrapeBatchActionResult(
        affected=len(runnable), skipped=len(tasks) - len(runnable)
    )


@router.post("/scrape-profile-batches/delete", response_model=InfluencerScrapeBatchActionResult)
def delete_scrape_batch(
    payload: InfluencerScrapeBatchRunRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """批量删除抓取任务（正在抓取中的跳过），不影响已入库达人。"""
    tasks = _batch_task_query(db, user, payload.batch, payload.platform).all()
    deletable = [t for t in tasks if t.status not in ("pending", "running")]
    for t in deletable:
        db.delete(t)
    db.commit()
    return InfluencerScrapeBatchActionResult(
        affected=len(deletable), skipped=len(tasks) - len(deletable)
    )


@router.get("/scrape-profile/{task_id}", response_model=InfluencerScrapeTaskOut)
def get_scrape_profile(
    task_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """轮询自动抓取任务状态/结果。"""
    task = db.get(InfluencerScrapeTask, task_id)
    if not task or (task.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="task not found")
    return _scrape_task_out(db, task)


@router.patch("/scrape-profile/{task_id}", response_model=InfluencerScrapeTaskOut)
def update_scrape_profile(
    task_id: int,
    payload: InfluencerScrapeTaskUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """修改暂存任务字段（目前支持改平台，纠正批量导入时选错的平台）。"""
    task = db.get(InfluencerScrapeTask, task_id)
    if not task or (task.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="task not found")
    if payload.platform is not None:
        platform = payload.platform.strip().lower()
        if platform not in KNOWN_PLATFORMS:
            raise HTTPException(status_code=400, detail="不支持的抓取平台")
        task.platform = platform
    db.commit()
    db.refresh(task)
    return _scrape_task_out(db, task)


@router.delete("/scrape-profile/{task_id}", response_model=Msg)
def delete_scrape_profile(
    task_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """删除一条暂存/抓取任务（不影响已入库的达人）。"""
    task = db.get(InfluencerScrapeTask, task_id)
    if not task or (task.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="task not found")
    db.delete(task)
    db.commit()
    return Msg(msg="deleted")


@router.post("/scrape-profile/{task_id}/run", response_model=InfluencerScrapeTaskOut)
def run_scrape_profile(
    task_id: int,
    payload: InfluencerScrapeRunRequest | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """对暂存/失败的任务发起（重新）抓取：置为 pending 并后台跑对应 actor。

    payload.auto_save=True 时抓完直接入库，save_status 指定入库后的建联状态。
    """
    task = db.get(InfluencerScrapeTask, task_id)
    if not task or (task.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="task not found")
    if task.status in ("running", "pending"):
        raise HTTPException(status_code=400, detail="该任务正在抓取中")
    if (task.platform or "facebook") not in SCRAPABLE_PLATFORMS:
        raise HTTPException(
            status_code=400, detail=f"平台 {task.platform} 暂不支持自动抓取资料"
        )
    task.status = "pending"
    task.error = None
    task.result = None
    task.started_at = None
    task.finished_at = None
    auto_save = bool(payload and payload.auto_save)
    status = _parse_save_status(payload.save_status if payload else None)
    db.commit()
    db.refresh(task)
    threading.Thread(
        target=_run_scrape_profile_bg,
        args=(task.id, auto_save, status.value if status else None),
        daemon=True,
    ).start()
    return _scrape_task_out(db, task)


@router.post("/scrape-profile/{task_id}/save", response_model=InfluencerScrapeSaveResult)
def save_scrape_profile(
    task_id: int,
    payload: InfluencerScrapeTaskSaveRequest | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """把某个已完成的抓取任务结果存入建联达人库（按主页ID/邮箱去重）。

    命中已有达人时复用、不重复创建，并通过 created=False 告知前端。
    """
    task = db.get(InfluencerScrapeTask, task_id)
    if not task or (task.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="task not found")
    if task.status != "done" or not isinstance(task.result, dict) or not task.result:
        raise HTTPException(status_code=400, detail="该任务尚未抓取完成，无法存入")
    notes = payload.notes if payload else None
    if (task.platform or "facebook") == "instagram":
        inf, created = influencer_service.create_influencer_from_ig_form(
            db, owner_id=task.owner_id, form=task.result, notes=notes
        )
    else:
        inf, created = influencer_service.create_influencer_from_form(
            db, owner_id=task.owner_id, form=task.result, notes=notes
        )
    _bind_platform_by_name(db, inf, task.platform)
    return InfluencerScrapeSaveResult(influencer=InfluencerOut.model_validate(inf), created=created)


@router.get("/{iid}", response_model=InfluencerDetailOut)
def get_influencer(
    iid: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    inf = db.get(Influencer, iid)
    if not inf or (inf.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="influencer not found")
    socials = (
        db.query(InfluencerSocialAccount)
        .filter(InfluencerSocialAccount.influencer_id == iid)
        .all()
    )
    post_ids = [
        p.id for p in db.query(Post.id).filter(Post.influencer_id == iid).all()
    ]
    influencer_service.link_outreach_logs_for_influencer(db, inf)
    has_outreach = (
        db.query(DmOutreachLog.id)
        .filter(DmOutreachLog.influencer_id == iid)
        .first()
        is not None
    )
    inf.has_outreach = has_outreach
    out = InfluencerDetailOut.model_validate(inf)
    out.social_accounts = [SocialAccountOut.model_validate(s) for s in socials]
    out.source_post_ids = post_ids
    return out


@router.get("/{iid}/outreach-logs", response_model=list[DmOutreachLogOut])
def list_influencer_outreach_logs(
    iid: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """该达人的私信建联发送记录（各看各的：仅返回当前用户发送的记录）。"""
    inf = db.get(Influencer, iid)
    if not inf or (inf.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="influencer not found")
    influencer_service.link_outreach_logs_for_influencer(db, inf)
    rows = (
        db.query(DmOutreachLog)
        .filter(
            DmOutreachLog.influencer_id == iid,
            DmOutreachLog.owner_id == user.id,
        )
        .order_by(DmOutreachLog.id.desc())
        .all()
    )
    return rows


@router.get("/{iid}/posts")
def list_influencer_posts(
    iid: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """返回该达人所有来源帖子（含 AI 评分、原帖链接）。"""
    inf = db.get(Influencer, iid)
    if not inf or (inf.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="influencer not found")
    rows = (
        db.query(Post)
        .filter(Post.influencer_id == iid)
        .order_by(Post.id.desc())
        .all()
    )
    return [
        {
            "id": p.id,
            "task_id": p.task_id,
            "url": p.url,
            "text": p.text,
            "author_name": p.author_name,
            "likes": p.likes,
            "comments_count": p.comments_count,
            "shares": p.shares,
            "ai_passed": p.ai_passed,
            "ai_score": p.ai_score,
            "ai_reason": p.ai_reason,
            "published_at": p.published_at,
        }
        for p in rows
    ]


@router.put("/{iid}", response_model=InfluencerOut)
def update_influencer(
    iid: int,
    payload: InfluencerUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    inf = db.get(Influencer, iid)
    if not inf or (inf.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="influencer not found")
    data = payload.model_dump(exclude_unset=True)
    _ensure_platform_access(db, inf.owner_id, data.get("platform_id"))
    if "country_id" in data:
        country = _country_or_400(db, data["country_id"])
        # 兼容字段跟着关联国家走，导出/旧筛选才对得上
        data["country"] = country.code if country else None
    for k, v in data.items():
        setattr(inf, k, v)
    db.commit()
    db.refresh(inf)
    return inf


@router.delete("/{iid}", response_model=Msg)
def delete_influencer(
    iid: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    inf = db.get(Influencer, iid)
    if not inf or (inf.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="influencer not found")
    db.delete(inf)
    db.commit()
    return Msg()


@router.post("/from-scrape", response_model=InfluencerOut)
def create_from_scrape(
    payload: InfluencerFromScrapeRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """从抓取的【待审核博主】列表点击【建联】入库。"""
    post: Post | None = None
    if payload.post_id:
        post = db.get(Post, payload.post_id)
        if not post or (post.owner_id != user.id and not is_admin(user)):
            raise HTTPException(status_code=404, detail="post not found")
    elif payload.author_url:
        post = (
            db.query(Post)
            .filter(Post.owner_id == user.id, Post.author_url == payload.author_url)
            .order_by(Post.id.desc())
            .first()
        )

    # page_profile 中可能带 _source_post_ids（来自抓取 step4 时回写的源帖子）
    src_ids = list(payload.source_post_ids or [])
    if payload.page_profile and isinstance(payload.page_profile, dict):
        extra_ids = payload.page_profile.get("_source_post_ids")
        if isinstance(extra_ids, list):
            src_ids.extend(int(x) for x in extra_ids if x)

    inf = influencer_service.create_from_scrape(
        db=db,
        owner_id=user.id,
        post=post,
        page_profile=payload.page_profile,
        notes=payload.notes,
        source_post_ids=src_ids or None,
    )
    return inf


@router.post("/{iid}/social-accounts", response_model=SocialAccountOut)
def add_social_account(
    iid: int,
    payload: SocialAccountCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    inf = db.get(Influencer, iid)
    if not inf or (inf.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="influencer not found")
    fields = payload.model_dump()
    if fields.get("platform_id") is None:
        fields["platform_id"] = influencer_service.resolve_platform_id(db, payload.platform)
    _ensure_platform_access(db, inf.owner_id, fields.get("platform_id"))
    if not fields.get("handle"):
        fields["handle"] = influencer_service.handle_from_url(
            fields.get("url"), fields.get("page_id")
        )
    sa = InfluencerSocialAccount(influencer_id=iid, **fields)
    db.add(sa)
    db.commit()
    db.refresh(sa)
    return sa


@router.post("/{iid}/social-accounts/scrape", response_model=InfluencerScrapeTaskOut)
def scrape_social_account(
    iid: int,
    payload: SocialAccountScrapeRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """列表「一键抓取」：选中该达人的某条关联账号（仅 Facebook / Instagram）后后台抓取，
    结果自动回写到这条账号（主表只补空的「人」维度字段）。"""
    inf = db.get(Influencer, iid)
    if not inf or (inf.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="influencer not found")
    sa = db.get(InfluencerSocialAccount, payload.social_account_id)
    if not sa or sa.influencer_id != iid:
        raise HTTPException(status_code=404, detail="social account not found")
    platform = sa.platform.value if isinstance(sa.platform, SocialPlatform) else str(sa.platform)
    if platform not in SCRAPABLE_PLATFORMS:
        raise HTTPException(status_code=400, detail=f"平台 {platform} 暂不支持一键抓取，目前仅支持 Facebook / Instagram")
    url = (sa.url or "").strip()
    if not url and platform == "instagram" and sa.handle:
        url = influencer_service.build_ig_profile_url(sa.handle)
    if not url:
        raise HTTPException(status_code=400, detail="该账号没有主页链接，请先编辑账号补上链接")
    task = InfluencerScrapeTask(
        owner_id=inf.owner_id,
        platform=platform,
        url=url,
        status="pending",
        influencer_id=iid,
        social_account_id=sa.id,
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    threading.Thread(
        target=_run_scrape_profile_bg, args=(task.id, True), daemon=True
    ).start()
    return _scrape_task_out(db, task)


@router.put("/{iid}/social-accounts/{sid}", response_model=SocialAccountOut)
def update_social_account(
    iid: int,
    sid: int,
    payload: SocialAccountUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """编辑社交账号：平台关联挂在账号上，改平台只影响这个账号。"""
    inf = db.get(Influencer, iid)
    if not inf or (inf.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="influencer not found")
    sa = db.get(InfluencerSocialAccount, sid)
    if not sa or sa.influencer_id != iid:
        raise HTTPException(status_code=404, detail="social account not found")
    data = payload.model_dump(exclude_unset=True)
    if "platform_id" in data:
        _ensure_platform_access(db, inf.owner_id, data["platform_id"])
    for field, value in data.items():
        setattr(sa, field, value)
    if data.get("platform") and "platform_id" not in data:
        sa.platform_id = influencer_service.resolve_platform_id(db, sa.platform)
    if not sa.handle:
        sa.handle = influencer_service.handle_from_url(sa.url, sa.page_id)
    db.commit()
    db.refresh(sa)
    return sa


@router.delete("/{iid}/social-accounts/{sid}", response_model=Msg)
def delete_social_account(
    iid: int,
    sid: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    inf = db.get(Influencer, iid)
    if not inf or (inf.owner_id != user.id and not is_admin(user)):
        raise HTTPException(status_code=404, detail="influencer not found")
    sa = db.get(InfluencerSocialAccount, sid)
    if not sa or sa.influencer_id != iid:
        raise HTTPException(status_code=404, detail="social account not found")
    db.delete(sa)
    db.commit()
    return Msg()
